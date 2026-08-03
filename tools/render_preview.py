"""
샘플 워크북을 HTML 표로 재현한다. 스크린샷 대신 쓴다.

실계정 리포트는 게시물 본문과 성과가 노출되므로 캡처할 수 없다. 동봉 샘플은
합성 데이터라 노출될 것이 없고, 셀 서식(헤더 색·강조 fill·bold)을 openpyxl로
읽어 그대로 옮기면 이미지 없이 랜딩페이지에 실물을 보여줄 수 있다.

    python3 tools/render_preview.py               # 기본 샘플 -> docs/preview.html
    python3 tools/render_preview.py -i a.xlsx -o b.html
"""

import argparse
import html
import os
import sys

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# (시트명, 화면 제목, 시작행, 행 수, 열 수). 시작행 None이면 판정이 갈리는 구간을 찾는다.
VIEWS = [
    ("Time of Day", "Time of Day", None, 9, 8),
    ("Content Optimization", "Content Optimization", 3, 10, 6),
    ("Viral Deep Dive", "Viral Deep Dive — view concentration", "6.", 11, 3),
]

# openpyxl이 주는 ARGB를 CSS로. 워크북이 쓰는 색만 매핑하고 나머지는 그대로 쓴다.
KNOWN = {
    "002F5496": ("#2f5496", "#ffffff"),   # 헤더
    "00FFF2CC": ("#fff2cc", "#1a1a1a"),   # 강조
    "00E2EFDA": ("#e2efda", "#1a1a1a"),   # 양호
    "00FCE4EC": ("#fce4ec", "#1a1a1a"),   # 주의
}


def cell_style(cell):
    fill = cell.fill
    rgb = fill.start_color.rgb if fill and fill.fill_type else None
    styles = []
    if isinstance(rgb, str) and rgb not in ("00000000", "FFFFFFFF"):
        bg, fg = KNOWN.get(rgb, ("#" + rgb[-6:], "#1a1a1a"))
        styles.append(f"background:{bg}")
        styles.append(f"color:{fg}")
    if cell.font and cell.font.bold:
        styles.append("font-weight:700")
    if isinstance(cell.value, (int, float)):
        styles.append("text-align:right")
        styles.append("font-variant-numeric:tabular-nums")
    return ";".join(styles)


def find_row(ws, prefix):
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row and isinstance(row[0], str) and row[0].startswith(prefix):
            return i
    return 1


def pick_start(ws, start):
    """판정이 갈리는 구간을 골라야 표가 뭔가를 말해준다."""
    if isinstance(start, int):
        return start
    if isinstance(start, str):
        return find_row(ws, start)
    interesting = [
        i for i, row in enumerate(ws.iter_rows(min_row=3, max_row=26, values_only=True), 3)
        if len(row) > 7 and isinstance(row[7], str) and ("🟢" in row[7] or "🔴" in row[7])
    ]
    return min(interesting) if interesting else 3


def render_sheet(ws, title, start, count, maxcol):
    header = [c for c in ws[2][:maxcol]]
    rows = list(ws.iter_rows(min_row=start, max_row=start + count - 1, max_col=maxcol))
    out = ['<figure class="xl">', f"<figcaption>{html.escape(title)}</figcaption>",
           '<div class="xl-scroll"><table>']
    if start > 2:  # 헤더 행을 따로 붙인다 (섹션 중간부터 자를 때)
        out.append("<thead><tr>" + "".join(
            f'<th style="{cell_style(c)}">{html.escape(str(c.value or ""))}</th>'
            for c in header) + "</tr></thead>")
    out.append("<tbody>")
    for row in rows:
        cells = [c for c in row]
        while cells and cells[-1].value is None:
            cells.pop()
        if not cells:
            continue
        out.append("<tr>" + "".join(
            f'<td style="{cell_style(c)}">{html.escape("" if c.value is None else str(c.value))}</td>'
            for c in cells) + "</tr>")
    out.append("</tbody></table></div></figure>")
    return "\n".join(out)


CSS = """<style>
.xl{margin:18px 0;border:1px solid var(--border,#3a3a3a);border-radius:10px;overflow:hidden;
    background:var(--card,#242424)}
.xl figcaption{padding:11px 14px;font-weight:600;font-size:.9rem;
    border-bottom:1px solid var(--border,#3a3a3a);color:var(--muted,#b8b8b8)}
.xl-scroll{overflow-x:auto;-webkit-overflow-scrolling:touch}
.xl table{border-collapse:collapse;font-size:.85rem;width:100%;min-width:420px}
.xl th,.xl td{padding:7px 11px;border-bottom:1px solid var(--border,#3a3a3a);
    white-space:nowrap;text-align:left}
.xl tbody tr:last-child td{border-bottom:none}
</style>"""


def main(argv=None):
    parser = argparse.ArgumentParser(description="샘플 워크북 → HTML 프리뷰")
    parser.add_argument("-i", "--input",
                        default=os.path.join(ROOT, "output", "sample_report_en.xlsx"))
    parser.add_argument("-o", "--output",
                        default=os.path.join(ROOT, "docs", "preview.html"))
    args = parser.parse_args(argv)

    if not os.path.isfile(args.input):
        print(f"[ERROR] 입력 없음: {args.input}")
        print("  먼저: python3 export_excel.py -i samples/sample_analysis.json "
              "--lang en -o output/sample_report_en.xlsx")
        sys.exit(1)

    wb = openpyxl.load_workbook(args.input)
    parts = [CSS]
    for sheet, title, start, count, maxcol in VIEWS:
        if sheet not in wb.sheetnames:
            print(f"  [건너뜀] 시트 없음: {sheet}")
            continue
        ws = wb[sheet]
        parts.append(render_sheet(ws, title, pick_start(ws, start), count, maxcol))

    os.makedirs(os.path.dirname(args.output), exist_ok=True)
    with open(args.output, "w", encoding="utf-8") as f:
        f.write("\n".join(parts) + "\n")
    print(f"프리뷰 생성: {args.output}")


if __name__ == "__main__":
    main()
