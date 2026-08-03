import json
import os
import sys
from datetime import datetime, timedelta, timezone
from calendar import monthrange
from collections import Counter

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import i18n

JST = timezone(timedelta(hours=9))

# Locale-independent weekday names (strftime %A is locale-dependent → KeyError on Korean/Japanese systems)
_WEEKDAY_NAMES = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]


def _weekday_en(dt):
    """Return English weekday name regardless of system locale."""
    return _WEEKDAY_NAMES[dt.weekday()]
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
SUB_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
HIGHLIGHT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

VIRAL_PERCENTILE = 99
# 조회수가 이 값 미만이면 '데드' 게시물로 간주 (여러 시트가 동일 기준을 공유)
DEAD_VIEWS = 500


def style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def auto_width(ws, min_width=10, max_width=60):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        length = max((len(str(c.value or "")) for c in col_cells), default=0)
        ws.column_dimensions[col_letter].width = min(max(length + 2, min_width), max_width)


def parse_ts(ts_str):
    if not ts_str:
        return None
    try:
        return datetime.fromisoformat(ts_str.replace("+0000", "+00:00")).astimezone(JST)
    except ValueError:
        return None


def load_data(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def sheet_all_posts(wb, posts):
    ws = wb.create_sheet("전체 게시물")
    headers = [
        "번호", "작성일시(JST)", "미디어타입", "내용(100자)",
        "조회수", "좋아요", "답글", "리포스트", "인용", "공유",
        "인게이지먼트", "좋아요율(%)", "링크"
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    sorted_posts = sorted(posts, key=lambda p: p.get("timestamp") or "", reverse=True)

    for i, p in enumerate(sorted_posts, 1):
        ins = (p.get("insights") or {})
        views = ins.get("views", 0)
        likes = ins.get("likes", 0)
        replies = ins.get("replies", 0)
        reposts = ins.get("reposts", 0)
        quotes = ins.get("quotes", 0)
        shares = ins.get("shares", 0)
        engagement = likes + replies + reposts + quotes
        like_rate = (likes / views * 100) if views > 0 else 0

        dt = parse_ts(p.get("timestamp"))
        dt_str = dt.strftime("%Y-%m-%d %H:%M") if dt else ""
        text = (p.get("text") or "").replace("\n", " ")[:100]

        ws.append([
            i, dt_str, p.get("media_type", ""), text,
            views, likes, replies, reposts, quotes, shares,
            engagement, round(like_rate, 2), p.get("permalink", "")
        ])

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    auto_width(ws)


def sheet_ranking(wb, posts, followers):
    ws = wb.create_sheet("인사이트 랭킹")

    def add_section(title, sorted_list, start_row):
        ws.cell(row=start_row, column=1, value=title).font = Font(bold=True, size=13)
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=10)

        headers = ["순위", "내용(50자)", "작성일", "조회수", "좋아요", "답글", "리포스트", "인게이지먼트", "좋아요율(%)", "링크"]
        for c, h in enumerate(headers, 1):
            ws.cell(row=start_row + 1, column=c, value=h)
        style_header(ws, start_row + 1, len(headers))

        for i, p in enumerate(sorted_list[:30], 1):
            ins = (p.get("insights") or {})
            views = ins.get("views", 0)
            likes = ins.get("likes", 0)
            replies = ins.get("replies", 0)
            reposts = ins.get("reposts", 0)
            engagement = likes + replies + reposts + ins.get("quotes", 0)
            like_rate = (likes / views * 100) if views > 0 else 0
            dt = parse_ts(p.get("timestamp"))
            dt_str = dt.strftime("%Y-%m-%d") if dt else ""
            text = (p.get("text") or "").replace("\n", " ")[:50]
            row = start_row + 2 + i - 1
            ws.append([])
            for c, v in enumerate([i, text, dt_str, views, likes, replies, reposts, engagement, round(like_rate, 2), p.get("permalink", "")], 1):
                ws.cell(row=row, column=c, value=v)
            if i <= 3:
                for c in range(1, len(headers) + 1):
                    ws.cell(row=row, column=c).fill = HIGHLIGHT_FILL

        return start_row + 2 + min(len(sorted_list), 30) + 2

    active_posts = [p for p in posts if p.get("media_type") != "REPOST_FACADE"]

    by_views = sorted(active_posts, key=lambda p: (p.get("insights") or {}).get("views", 0), reverse=True)
    row = add_section("📊 조회수 TOP 30", by_views, 1)

    by_engagement = sorted(active_posts, key=lambda p: sum([
        (p.get("insights") or {}).get("likes", 0),
        (p.get("insights") or {}).get("replies", 0),
        (p.get("insights") or {}).get("reposts", 0),
        (p.get("insights") or {}).get("quotes", 0),
    ]), reverse=True)
    row = add_section("❤️ 인게이지먼트 TOP 30", by_engagement, row)

    min_views = 500
    qualified = [p for p in active_posts if (p.get("insights") or {}).get("views", 0) >= min_views]
    by_like_rate = sorted(qualified, key=lambda p: (p.get("insights") or {}).get("likes", 0) / max((p.get("insights") or {}).get("views", 1), 1), reverse=True)
    row = add_section(f"🎯 좋아요율 TOP 30 (조회 {min_views}+ 기준)", by_like_rate, row)

    by_viral = sorted(active_posts, key=lambda p: (p.get("insights") or {}).get("reposts", 0) + (p.get("insights") or {}).get("quotes", 0), reverse=True)
    add_section("🔄 바이럴(리포스트+인용) TOP 30", by_viral, row)

    auto_width(ws)


_TIME_MIN_SAMPLE = 30


def _time_bucket_row(label, metrics, overall_median):
    """구간 1행 = (셀 리스트, 우수 여부). 평균은 바이럴 1건에 끌려가므로 판정은 중앙값 기준."""
    n = len(metrics)
    if not n:
        return [label, 0, 0, 0, 0.0, 0.0, 0.0, "표본부족"], False
    views = [m["views"] for m in metrics]
    median_views = _median(views)
    dead_rate = sum(1 for v in views if v < DEAD_VIEWS) / n * 100
    if n < _TIME_MIN_SAMPLE:
        verdict = "표본부족"
    elif median_views >= overall_median * 1.2:
        verdict = "🟢우수"
    elif median_views <= overall_median * 0.8:
        verdict = "🔴회피"
    else:
        verdict = "⚪보통"
    return [
        label,
        n,
        round(median_views),
        round(_avg(views)),
        round(dead_rate, 2),
        round(_median([m["likes"] for m in metrics]), 1),
        round(_avg([m["engagement"] for m in metrics]), 1),
        verdict,
    ], verdict == "🟢우수"


def sheet_time_analysis(wb, posts):
    ws = wb.create_sheet("시간대 분석")
    annotated = [_post_metrics(p) for p in _active_posts(posts)]
    dated = [m for m in annotated if m["dt"]]

    hour_data = {}
    weekday_data = {}
    day_kr = {"Monday": "월", "Tuesday": "화", "Wednesday": "수", "Thursday": "목", "Friday": "금", "Saturday": "토", "Sunday": "일"}

    for m in dated:
        hour_data.setdefault(m["dt"].hour, []).append(m)
        weekday_data.setdefault(_weekday_en(m["dt"]), []).append(m)

    overall_median = _median([m["views"] for m in dated])

    ws.cell(row=1, column=1, value=f"시간대별 성과 (JST, 중앙값 기준 · 데드={DEAD_VIEWS}조회 미만 · 판정은 {_TIME_MIN_SAMPLE}건 이상)").font = Font(bold=True, size=13)
    headers = ["시간", "게시수", "중앙값조회", "평균조회", "데드율(%)", "중앙값좋아요", "평균인게이지먼트", "판정"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=2, column=c, value=h)
    style_header(ws, 2, len(headers))

    if not dated:
        ws.cell(row=3, column=1, value="데이터 없음")
        auto_width(ws)
        return

    for h in range(24):
        values, is_best = _time_bucket_row(f"{h:02d}:00", hour_data.get(h, []), overall_median)
        for c, v in enumerate(values, 1):
            ws.cell(row=3 + h, column=c, value=v)
        if is_best:
            _fill_row(ws, 3 + h, len(headers), HIGHLIGHT_FILL)

    row_start = 28
    ws.cell(row=row_start, column=1, value="요일별 성과").font = Font(bold=True, size=13)
    weekday_headers = ["요일", "게시수", "중앙값조회", "평균조회", "데드율(%)", "중앙값좋아요", "평균인게이지먼트", "판정"]
    for c, h in enumerate(weekday_headers, 1):
        ws.cell(row=row_start + 1, column=c, value=h)
    style_header(ws, row_start + 1, len(weekday_headers))

    for i, day in enumerate(["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]):
        values, is_best = _time_bucket_row(day_kr.get(day, day), weekday_data.get(day, []), overall_median)
        r = row_start + 2 + i
        for c, v in enumerate(values, 1):
            ws.cell(row=r, column=c, value=v)
        if is_best:
            _fill_row(ws, r, len(weekday_headers), HIGHLIGHT_FILL)

    auto_width(ws)


def sheet_demographics(wb, demographics, followers):
    ws = wb.create_sheet("팔로워 인구통계")
    label_map = {"country": "국가", "city": "도시", "gender": "성별", "age": "연령대"}
    row = 1

    ws.cell(row=row, column=1, value=f"총 팔로워: {followers:,}명" if followers else "총 팔로워: -").font = Font(bold=True, size=13)
    row += 2

    for key in ["country", "city", "gender", "age"]:
        data = demographics.get(key, {})
        if not data or not isinstance(data, dict):
            continue
        ws.cell(row=row, column=1, value=label_map.get(key, key)).font = Font(bold=True, size=12)
        row += 1
        headers = [label_map.get(key, key), "팔로워수", "비율(%)"]
        for c, h in enumerate(headers, 1):
            ws.cell(row=row, column=c, value=h)
        style_header(ws, row, len(headers))
        row += 1

        total = sum(data.values())
        for k, v in sorted(data.items(), key=lambda x: x[1], reverse=True):
            pct = (v / total * 100) if total > 0 else 0
            ws.cell(row=row, column=1, value=k)
            ws.cell(row=row, column=2, value=v)
            ws.cell(row=row, column=3, value=round(pct, 1))
            row += 1
        row += 2

    auto_width(ws)


def _build_growth_strategies(active, type_stats, viral_posts, demographics, total_views, total_likes, total_replies, total_reposts, total_quotes):
    """Data-driven strategy bullets — no account-specific hardcoding."""
    strategies = []
    day_kr = {"Monday": "월", "Tuesday": "화", "Wednesday": "수", "Thursday": "목", "Friday": "금", "Saturday": "토", "Sunday": "일"}

    carousel = type_stats.get("CAROUSEL_ALBUM")
    text = type_stats.get("TEXT_POST")
    if carousel and text and carousel["count"] and text["count"]:
        c_avg = carousel["engagement"] / carousel["count"]
        t_avg = text["engagement"] / text["count"]
        if t_avg > 0 and c_avg > t_avg:
            strategies.append(
                f"캐러셀 활용 확대 — 평균 인게이지먼트 캐러셀 {c_avg:.1f} vs 텍스트 {t_avg:.1f} "
                f"({c_avg / t_avg:.1f}배). 현재 {carousel['count']}개 → 주 1-2회 목표"
            )
        elif c_avg > 0:
            strategies.append(
                f"미디어 타입 실험 — 캐러셀 {c_avg:.1f} / 텍스트 {t_avg:.1f} 평균 인게이지먼트. 상위 타입 비중 확대"
            )

    hour_views = Counter()
    hour_count = Counter()
    for p in active:
        dt = parse_ts(p.get("timestamp"))
        if not dt:
            continue
        hour_views[dt.hour] += (p.get("insights") or {}).get("views", 0)
        hour_count[dt.hour] += 1
    if hour_count:
        avg_by_hour = {h: hour_views[h] / hour_count[h] for h in hour_count}
        top_hours = sorted(avg_by_hour.items(), key=lambda x: x[1], reverse=True)[:3]
        hour_str = ", ".join(f"{h:02d}시" for h, _ in top_hours)
        strategies.append(f"골든타임 집중 — 평균 조회 상위 시간대(JST): {hour_str}. 핵심 콘텐츠를 이 시간대에 게시")

    if viral_posts:
        viral_types = Counter(p.get("media_type") for p in viral_posts)
        top_type, top_n = viral_types.most_common(1)[0]
        strategies.append(
            f"바이럴 패턴 반복 — 상위 조회 바이럴 {len(viral_posts)}개 중 최빈 타입 {top_type}({top_n}건). "
            "해당 포맷·주제 패턴 재사용"
        )

    strategies.append(
        "좋아요율 높은 주제 강화 — 좋아요율 TOP 게시물 패턴을 분석해 반복 (좋아요율 = 팬 충성도 지표)"
    )

    demo_parts = []
    for key, label in [("age", "연령"), ("gender", "성별"), ("country", "국가")]:
        data = demographics.get(key) or {}
        if isinstance(data, dict) and data:
            top_k, _ = max(data.items(), key=lambda x: x[1])
            demo_parts.append(f"{label} {top_k}")
    if demo_parts:
        strategies.append(
            f"팔로워 인구통계 활용 — 핵심층: {', '.join(demo_parts)}. 이 타겟에 맞는 콘텐츠 톤 유지"
        )

    eng_rate = (total_likes + total_replies + total_reposts + total_quotes) / max(total_views, 1) * 100
    strategies.append(
        f"답글 유도 — 현재 인게이지먼트율 {eng_rate:.2f}%. 질문형 마무리로 답글 비중을 높이면 알고리즘 가중치에 유리"
    )

    weekday_count = Counter()
    for p in active:
        dt = parse_ts(p.get("timestamp"))
        if dt:
            weekday_count[_weekday_en(dt)] += 1
    if weekday_count:
        weekend = weekday_count.get("Saturday", 0) + weekday_count.get("Sunday", 0)
        weekday = sum(weekday_count.get(d, 0) for d in ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"])
        if weekday > 0 and weekend < weekday * 0.35:
            strategies.append(
                "주말 활동 강화 — 토·일 게시량이 평일 대비 적음. 경쟁 콘텐츠가 줄어 도달 기회일 수 있음"
            )
        elif weekend > 0:
            top_day = max(weekday_count.items(), key=lambda x: x[1])[0]
            strategies.append(
                f"게시 요일 분산 — 현재 최빈 요일 {day_kr.get(top_day, top_day)}. "
                "성과 상위 요일에 핵심 콘텐츠 배치"
            )

    return strategies


def sheet_insights_report(wb, posts, user_insights, followers, username="", demographics=None):
    ws = wb.create_sheet("성장 인사이트")
    active = [p for p in posts if p.get("media_type") != "REPOST_FACADE"]
    demographics = demographics or {}

    total_views = sum((p.get("insights") or {}).get("views", 0) for p in active)
    total_likes = sum((p.get("insights") or {}).get("likes", 0) for p in active)
    total_replies = sum((p.get("insights") or {}).get("replies", 0) for p in active)
    total_reposts = sum((p.get("insights") or {}).get("reposts", 0) for p in active)
    total_quotes = sum((p.get("insights") or {}).get("quotes", 0) for p in active)

    type_stats = {}
    for p in active:
        mt = p.get("media_type", "UNKNOWN")
        ins = (p.get("insights") or {})
        if mt not in type_stats:
            type_stats[mt] = {"count": 0, "views": 0, "likes": 0, "engagement": 0}
        type_stats[mt]["count"] += 1
        type_stats[mt]["views"] += ins.get("views", 0)
        eng = ins.get("likes", 0) + ins.get("replies", 0) + ins.get("reposts", 0) + ins.get("quotes", 0)
        type_stats[mt]["engagement"] += eng

    viral_min = _viral_threshold([(p, _post_metrics(p)) for p in active])
    viral_posts = [p for p in active if (p.get("insights") or {}).get("views", 0) >= viral_min]
    high_engage = [p for p in active if (p.get("insights") or {}).get("views", 0) >= DEAD_VIEWS]
    high_engage_sorted = sorted(high_engage,
        key=lambda p: ((p.get("insights") or {}).get("likes", 0) + (p.get("insights") or {}).get("replies", 0)) / max((p.get("insights") or {}).get("views", 1), 1),
        reverse=True)

    row = 1
    ws.cell(row=row, column=1, value=f"@{username} 성장 인사이트 리포트").font = Font(bold=True, size=14)
    ws.cell(row=row + 1, column=1, value=f"분석일: {datetime.now(JST).strftime('%Y-%m-%d')} / 팔로워: {f'{followers:,}명' if followers else '-'} / 게시물: {len(active)}개 (리포스트 제외)")
    row += 3

    ws.cell(row=row, column=1, value="1. 핵심 지표").font = Font(bold=True, size=12)
    row += 1
    eng_rate = (total_likes + total_replies + total_reposts + total_quotes) / max(total_views, 1) * 100
    kpi = [
        ["총 조회수", f"{total_views:,}"],
        ["총 좋아요", f"{total_likes:,}"],
        ["총 답글", f"{total_replies:,}"],
        ["총 리포스트", f"{total_reposts:,}"],
        ["평균 조회수/게시물", f"{total_views // max(len(active), 1):,}"],
        ["평균 좋아요/게시물", f"{total_likes / max(len(active), 1):.1f}"],
        ["전체 인게이지먼트율", f"{eng_rate:.2f}%"],
        [f"바이럴 게시물 ({viral_min:,}+조회)", f"{len(viral_posts)}개 ({len(viral_posts)/max(len(active), 1)*100:.1f}%)"],
        ["30일 조회수", f"{user_insights['30d_views']:,}" if user_insights.get("30d_views") else "-"],
    ]
    for label, val in kpi:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=val)
        row += 1
    row += 2

    ws.cell(row=row, column=1, value="2. 미디어 타입별 성과").font = Font(bold=True, size=12)
    row += 1
    headers = ["타입", "게시물수", "총조회수", "평균조회수", "총인게이지먼트", "평균인게이지먼트"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header(ws, row, len(headers))
    row += 1
    for mt, s in sorted(type_stats.items(), key=lambda x: x[1]["engagement"] / max(x[1]["count"], 1), reverse=True):
        cnt = s["count"]
        ws.cell(row=row, column=1, value=mt)
        ws.cell(row=row, column=2, value=cnt)
        ws.cell(row=row, column=3, value=s["views"])
        ws.cell(row=row, column=4, value=round(s["views"] / cnt))
        ws.cell(row=row, column=5, value=s["engagement"])
        ws.cell(row=row, column=6, value=round(s["engagement"] / cnt, 1))
        row += 1
    row += 2

    ws.cell(row=row, column=1, value="3. 좋아요율 TOP 20 (조회 500+ 기준) — 팬이 반응하는 콘텐츠").font = Font(bold=True, size=12)
    row += 1
    headers = ["순위", "내용(60자)", "조회수", "좋아요", "좋아요율(%)", "작성일"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header(ws, row, len(headers))
    row += 1
    for i, p in enumerate(high_engage_sorted[:20], 1):
        ins = (p.get("insights") or {})
        views = ins.get("views", 1)
        likes = ins.get("likes", 0)
        dt = parse_ts(p.get("timestamp"))
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=(p.get("text") or "").replace("\n", " ")[:60])
        ws.cell(row=row, column=3, value=views)
        ws.cell(row=row, column=4, value=likes)
        ws.cell(row=row, column=5, value=round(likes / max(views, 1) * 100, 2))
        ws.cell(row=row, column=6, value=dt.strftime("%Y-%m-%d") if dt else "")
        row += 1
    row += 2

    ws.cell(row=row, column=1, value=f"4. 바이럴 게시물 패턴 분석 ({viral_min:,}+ 조회)").font = Font(bold=True, size=12)
    row += 1
    if viral_posts:
        viral_types = Counter(p.get("media_type") for p in viral_posts)
        viral_hours = Counter()
        viral_weekdays = Counter()
        for p in viral_posts:
            dt = parse_ts(p.get("timestamp"))
            if dt:
                viral_hours[dt.hour] += 1
                viral_weekdays[_weekday_en(dt)] += 1
        top_hours = viral_hours.most_common(3)
        top_days = viral_weekdays.most_common(3)
        day_kr = {"Monday": "월", "Tuesday": "화", "Wednesday": "수", "Thursday": "목", "Friday": "금", "Saturday": "토", "Sunday": "일"}

        patterns = [
            ["바이럴 게시물 수", f"{len(viral_posts)}개"],
            ["주요 미디어 타입", ", ".join(f"{k}({v})" for k, v in viral_types.most_common())],
            ["바이럴 잘 되는 시간", ", ".join(f"{h}시({c}건)" for h, c in top_hours)],
            ["바이럴 잘 되는 요일", ", ".join(f"{day_kr.get(d,d)}({c}건)" for d, c in top_days)],
        ]
        for label, val in patterns:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=val)
            row += 1
    row += 2

    ws.cell(row=row, column=1, value="5. 성장 전략 제안").font = Font(bold=True, size=12)
    row += 1
    strategies = _build_growth_strategies(
        active, type_stats, viral_posts, demographics,
        total_views, total_likes, total_replies, total_reposts, total_quotes,
    )
    for s in strategies:
        ws.cell(row=row, column=1, value=f"• {s}")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
        row += 1

    auto_width(ws)


def _active_posts(posts):
    return [p for p in posts if p.get("media_type") != "REPOST_FACADE"]


def _post_metrics(post):
    ins = post.get("insights") or {}
    views = ins.get("views", 0)
    likes = ins.get("likes", 0)
    replies = ins.get("replies", 0)
    reposts = ins.get("reposts", 0)
    quotes = ins.get("quotes", 0)
    shares = ins.get("shares", 0)
    return {
        "views": views,
        "likes": likes,
        "replies": replies,
        "reposts": reposts,
        "quotes": quotes,
        "shares": shares,
        "engagement": likes + replies + reposts + quotes,
        "like_rate": (likes / views * 100) if views > 0 else 0,
        "dt": parse_ts(post.get("timestamp")),
        "text": (post.get("text") or "").replace("\n", " "),
        "media_type": post.get("media_type", ""),
        "permalink": post.get("permalink", ""),
    }


def _avg(values):
    return sum(values) / len(values) if values else 0


def _median(values):
    if not values:
        return 0
    sorted_values = sorted(values)
    mid = len(sorted_values) // 2
    if len(sorted_values) % 2 == 1:
        return sorted_values[mid]
    return (sorted_values[mid - 1] + sorted_values[mid]) / 2


def _percentile(sorted_values, q):
    # 최근접 순위(nearest-rank). numpy 미사용, 오름차순 리스트 전제
    if not sorted_values:
        return 0
    return sorted_values[min(len(sorted_values) - 1, int(len(sorted_values) * q))]


def _viral_threshold(annotated):
    # 계정이 커지면 기준도 같이 올라가되, 과거 기준선 10000 아래로는 내려가지 않음
    views = sorted(m["views"] for _, m in annotated)
    return max(_percentile(views, VIRAL_PERCENTILE / 100), 10000)


def _set_section_title(ws, row, title, end_col=10):
    ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=13)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    return row + 1


def _set_headers(ws, row, headers):
    for c, header in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=header)
    style_header(ws, row, len(headers))
    return row + 1


def _fill_row(ws, row, max_col, fill):
    for c in range(1, max_col + 1):
        ws.cell(row=row, column=c).fill = fill


def sheet_viral_analysis(wb, posts):
    ws = wb.create_sheet("바이럴 심층분석")
    active_posts = _active_posts(posts)
    annotated = [(p, _post_metrics(p)) for p in active_posts]
    viral_min = _viral_threshold(annotated)
    viral = [(p, m) for p, m in annotated if m["views"] >= viral_min]
    normal = [(p, m) for p, m in annotated if m["views"] < viral_min]
    total_views = sum(m["views"] for _, m in annotated)
    overall_viral_ratio = (len(viral) / len(annotated) * 100) if annotated else 0
    day_order = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    day_kr = {"Monday": "월", "Tuesday": "화", "Wednesday": "수", "Thursday": "목", "Friday": "금", "Saturday": "토", "Sunday": "일"}
    row = 1

    row = _set_section_title(ws, row, f"1. 바이럴 게시물 목록 ({viral_min:,}+ 조회)")
    headers = ["순위", "작성일", "미디어타입", "내용(80자)", "조회수", "좋아요", "답글", "리포스트", "좋아요율(%)", "링크"]
    row = _set_headers(ws, row, headers)
    for i, (post, m) in enumerate(sorted(viral, key=lambda x: x[1]["views"], reverse=True), 1):
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=m["dt"].strftime("%Y-%m-%d") if m["dt"] else "")
        ws.cell(row=row, column=3, value=m["media_type"])
        ws.cell(row=row, column=4, value=m["text"][:80])
        ws.cell(row=row, column=5, value=m["views"])
        ws.cell(row=row, column=6, value=m["likes"])
        ws.cell(row=row, column=7, value=m["replies"])
        ws.cell(row=row, column=8, value=m["reposts"])
        ws.cell(row=row, column=9, value=round(m["like_rate"], 2))
        ws.cell(row=row, column=10, value=m["permalink"])
        if i <= 3:
            _fill_row(ws, row, len(headers), HIGHLIGHT_FILL)
        row += 1
    if not viral:
        ws.cell(row=row, column=1, value="해당 없음")
        row += 1
    row += 1

    row = _set_section_title(ws, row, "2. 바이럴 vs 일반 비교", end_col=8)
    headers = ["구분", "게시물수", "평균조회", "평균좋아요", "평균답글", "평균리포스트", "평균좋아요율(%)", "전체조회비중(%)"]
    row = _set_headers(ws, row, headers)
    viral_label = f"바이럴({viral_min:,}+)"
    compare_groups = [(viral_label, viral), (f"일반(<{viral_min:,})", normal)]
    for label, group in compare_groups:
        views_sum = sum(m["views"] for _, m in group)
        likes = [m["likes"] for _, m in group]
        replies = [m["replies"] for _, m in group]
        reposts = [m["reposts"] for _, m in group]
        like_rates = [m["like_rate"] for _, m in group]
        ws.append([
            label,
            len(group),
            round(views_sum / len(group)) if group else 0,
            round(_avg(likes), 1),
            round(_avg(replies), 1),
            round(_avg(reposts), 1),
            round(_avg(like_rates), 2),
            round((views_sum / total_views * 100), 2) if total_views > 0 else 0,
        ])
        if label == viral_label:
            _fill_row(ws, row, len(headers), HIGHLIGHT_FILL)
        row += 1
    row += 1

    row = _set_section_title(ws, row, "3. 바이럴 시간대 히트맵", end_col=4)
    headers = ["시간", "바이럴수", "일반수", "바이럴비율(%)"]
    row = _set_headers(ws, row, headers)
    hour_stats = {h: {"viral": 0, "normal": 0} for h in range(24)}
    for _, m in viral:
        if m["dt"]:
            hour_stats[m["dt"].hour]["viral"] += 1
    for _, m in normal:
        if m["dt"]:
            hour_stats[m["dt"].hour]["normal"] += 1
    for hour in range(24):
        v = hour_stats[hour]["viral"]
        n = hour_stats[hour]["normal"]
        ratio = (v / (v + n) * 100) if (v + n) > 0 else 0
        ws.append([f"{hour:02d}시", v, n, round(ratio, 2)])
        if ratio > overall_viral_ratio:
            _fill_row(ws, row, len(headers), HIGHLIGHT_FILL)
        row += 1
    row += 1

    row = _set_section_title(ws, row, "4. 바이럴 요일분포", end_col=4)
    headers = ["요일", "바이럴수", "일반수", "바이럴비율(%)"]
    row = _set_headers(ws, row, headers)
    weekday_stats = {day: {"viral": 0, "normal": 0} for day in day_order}
    for _, m in viral:
        if m["dt"]:
            weekday_stats[_weekday_en(m["dt"])]["viral"] += 1
    for _, m in normal:
        if m["dt"]:
            weekday_stats[_weekday_en(m["dt"])]["normal"] += 1
    for day in day_order:
        v = weekday_stats[day]["viral"]
        n = weekday_stats[day]["normal"]
        ratio = (v / (v + n) * 100) if (v + n) > 0 else 0
        ws.append([day_kr[day], v, n, round(ratio, 2)])
        if ratio > overall_viral_ratio:
            _fill_row(ws, row, len(headers), HIGHLIGHT_FILL)
        row += 1
    row += 1

    row = _set_section_title(ws, row, "5. 바이럴 미디어타입", end_col=4)
    headers = ["타입", "바이럴수", "전체수", "바이럴비율(%)"]
    row = _set_headers(ws, row, headers)
    type_stats = {}
    for _, m in annotated:
        mt = m["media_type"]
        if mt not in type_stats:
            type_stats[mt] = {"viral": 0, "total": 0}
        type_stats[mt]["total"] += 1
        if m["views"] >= viral_min:
            type_stats[mt]["viral"] += 1
    for media_type, stats in sorted(type_stats.items(), key=lambda x: ((x[1]["viral"] / x[1]["total"]) if x[1]["total"] else 0, x[1]["total"]), reverse=True):
        ratio = (stats["viral"] / stats["total"] * 100) if stats["total"] else 0
        ws.append([media_type, stats["viral"], stats["total"], round(ratio, 2)])
        row += 1
    row += 1

    row = _set_section_title(ws, row, "6. 조회수 집중도", end_col=2)
    headers = ["지표", "값"]
    row = _set_headers(ws, row, headers)
    desc_views = sorted((m["views"] for _, m in annotated), reverse=True)
    if not desc_views:
        ws.append(["데이터 없음", 0])
        row += 1
    else:
        n = len(desc_views)
        asc_views = desc_views[::-1]
        mean_v = _avg(desc_views)
        median_v = _median(desc_views)
        shares = []
        for q in (0.01, 0.05, 0.10):
            k = max(1, int(n * q))
            shares.append((sum(desc_views[:k]) / total_views * 100) if total_views else 0)
        top_one = (desc_views[0] / total_views * 100) if total_views else 0
        mean_ex_top = _avg(desc_views[1:])
        stats_rows = [
            ("원본게시물수", n),
            ("평균조회", round(mean_v)),
            ("중앙값조회", round(median_v)),
            ("평균÷중앙값 배수", round(mean_v / median_v, 2) if median_v else 0),
            ("P50", round(median_v)),
            ("P75", _percentile(asc_views, 0.75)),
            ("P90", _percentile(asc_views, 0.90)),
            ("P99", _percentile(asc_views, 0.99)),
            ("바이럴기준(적용값)", viral_min),
            ("상위1% 조회비중(%)", round(shares[0], 2)),
            ("상위5% 조회비중(%)", round(shares[1], 2)),
            ("상위10% 조회비중(%)", round(shares[2], 2)),
            ("최다조회 1건 비중(%)", round(top_one, 2)),
            ("1위 제외 평균", round(mean_ex_top)),
        ]
        for label, value in stats_rows:
            ws.append([label, value])
            row += 1
        ws.append([
            "해석",
            f"상위 1%({max(1, int(n * 0.01))}개)가 전체 조회의 {shares[0]:.1f}%, 최다 1건만으로 {top_one:.1f}%를 차지합니다. "
            f"평균 {round(mean_v):,}은 중앙값 {round(median_v):,}의 {(mean_v / median_v):.1f}배이며, 1위 게시물을 빼면 평균이 {round(mean_ex_top):,}로 내려갑니다. "
            "다른 시트의 '평균' 지표는 소수 이상치가 만든 값이므로 중앙값과 함께 해석하세요."
            if median_v else "중앙값이 0이라 배수 비교가 불가합니다. 조회수 데이터가 충분한지 확인하세요.",
        ])
        row += 1

    ws.freeze_panes = "A3"
    auto_width(ws)


def sheet_like_rate_analysis(wb, posts):
    ws = wb.create_sheet("좋아요율 심층분석")
    active_posts = _active_posts(posts)
    qualified = [(p, _post_metrics(p)) for p in active_posts if (p.get("insights") or {}).get("views", 0) >= DEAD_VIEWS]
    row = 1

    row = _set_section_title(ws, row, "1. 좋아요율 분포", end_col=5)
    headers = ["구간", "게시물수", "비율(%)", "평균조회수", "평균좋아요수"]
    row = _set_headers(ws, row, headers)
    buckets = [
        ("0-1%", 0, 1),
        ("1-2%", 1, 2),
        ("2-3%", 2, 3),
        ("3-5%", 3, 5),
        ("5-10%", 5, 10),
        ("10%+", 10, None),
    ]
    total_qualified = len(qualified)
    for label, low, high in buckets:
        bucket_posts = []
        for post, m in qualified:
            rate = m["like_rate"]
            if high is None:
                matched = rate >= low
            else:
                matched = low <= rate < high
            if matched:
                bucket_posts.append((post, m))
        ws.append([
            label,
            len(bucket_posts),
            round((len(bucket_posts) / total_qualified * 100), 2) if total_qualified else 0,
            round(_avg([m["views"] for _, m in bucket_posts])) if bucket_posts else 0,
            round(_avg([m["likes"] for _, m in bucket_posts]), 1) if bucket_posts else 0,
        ])
        row += 1
    row += 1

    top_posts = sorted(qualified, key=lambda x: x[1]["like_rate"], reverse=True)[:30]
    row = _set_section_title(ws, row, "2. 좋아요율 TOP 30", end_col=9)
    headers = ["순위", "내용(80자)", "작성일", "미디어타입", "조회수", "좋아요", "답글", "좋아요율(%)", "링크"]
    row = _set_headers(ws, row, headers)
    for i, (post, m) in enumerate(top_posts, 1):
        ws.append([
            i,
            m["text"][:80],
            m["dt"].strftime("%Y-%m-%d") if m["dt"] else "",
            m["media_type"],
            m["views"],
            m["likes"],
            m["replies"],
            round(m["like_rate"], 2),
            m["permalink"],
        ])
        if i <= 3:
            _fill_row(ws, row, len(headers), HIGHLIGHT_FILL)
        row += 1
    row += 1

    worst_posts = sorted(qualified, key=lambda x: x[1]["like_rate"])[:30]
    row = _set_section_title(ws, row, "3. 좋아요율 WORST 30", end_col=9)
    row = _set_headers(ws, row, headers)
    for i, (post, m) in enumerate(worst_posts, 1):
        ws.append([
            i,
            m["text"][:80],
            m["dt"].strftime("%Y-%m-%d") if m["dt"] else "",
            m["media_type"],
            m["views"],
            m["likes"],
            m["replies"],
            round(m["like_rate"], 2),
            m["permalink"],
        ])
        row += 1
    row += 1

    row = _set_section_title(ws, row, "4. 미디어타입별 좋아요율", end_col=4)
    headers = ["타입", "게시물수", "평균좋아요율(%)", "중앙값좋아요율(%)"]
    row = _set_headers(ws, row, headers)
    type_stats = {}
    for _, m in qualified:
        mt = m["media_type"]
        if mt not in type_stats:
            type_stats[mt] = []
        type_stats[mt].append(m["like_rate"])
    for media_type, rates in sorted(type_stats.items(), key=lambda x: _avg(x[1]), reverse=True):
        ws.append([media_type, len(rates), round(_avg(rates), 2), round(_median(rates), 2)])
        row += 1

    ws.freeze_panes = "A3"
    auto_width(ws)


def sheet_content_optimization(wb, posts):
    ws = wb.create_sheet("콘텐츠 최적화")
    active_posts = _active_posts(posts)
    annotated = [(p, _post_metrics(p)) for p in active_posts]
    row = 1

    row = _set_section_title(ws, row, "1. 글자수 구간별 성과 (조회수 vs 인게이지먼트율)", end_col=8)
    dead_cutoff = DEAD_VIEWS
    all_views = sum(m["views"] for _, m in annotated)
    type_counts = {}
    for _, m in annotated:
        type_counts[m["media_type"] or "UNKNOWN"] = type_counts.get(m["media_type"] or "UNKNOWN", 0) + 1
    type_note = ", ".join(f"{k} {v}건" for k, v in sorted(type_counts.items(), key=lambda x: -x[1]))
    ws.cell(row=row, column=1, value=f"미디어타입 전체 포함: {type_note or '데이터 없음'} / 데드율 = 조회 {round(dead_cutoff)} 미만 비율")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
    row += 1
    headers = ["글자수구간", "게시수", "중앙값조회", "평균조회", "중앙값ER(%)", "데드율(%)", "총조회비중(%)", "판정"]
    row = _set_headers(ws, row, headers)
    length_buckets = [
        ("0-50자", 0, 50),
        ("50-100자", 50, 100),
        ("100-150자", 100, 150),
        ("150-200자", 150, 200),
        ("200-300자", 200, 300),
        ("300-400자", 300, 400),
        ("400-500자", 400, 500),
        ("500자+", 500, None),
    ]
    bucket_rows = []
    for label, low, high in length_buckets:
        bucket = []
        for post, m in annotated:
            text_len = len((post.get("text") or "").replace("\n", " ").strip())
            if high is None:
                matched = text_len >= low
            else:
                matched = low <= text_len < high
            if matched:
                bucket.append(m)
        # ER은 shares 포함 — _post_metrics['engagement']는 shares를 제외하므로 여기서 직접 계산
        ers = [
            ((m["likes"] + m["replies"] + m["reposts"] + m["quotes"] + m["shares"]) / m["views"] * 100)
            if m["views"] > 0 else 0
            for m in bucket
        ]
        bucket_views = [m["views"] for m in bucket]
        bucket_rows.append({
            "label": label,
            "count": len(bucket),
            "median_views": round(_median(bucket_views)),
            "avg_views": round(_avg(bucket_views)),
            "median_er": round(_median(ers), 2),
            "dead_rate": round(len([v for v in bucket_views if v < dead_cutoff]) / len(bucket) * 100, 2) if bucket else 0,
            "view_share": round(sum(bucket_views) / all_views * 100, 2) if all_views else 0,
        })
    eligible = [r for r in bucket_rows if r["count"] >= 30]
    top_labels = {r["label"] for r in sorted(eligible, key=lambda x: -x["median_views"])[:2]}
    best_median = max((r["median_views"] for r in eligible), default=0)
    if not [r for r in bucket_rows if r["count"] > 0]:
        ws.append(["데이터 없음", 0, 0, 0, 0, 0, 0, "-"])
        row += 1
    else:
        for bucket in bucket_rows:
            if bucket["count"] < 30:
                verdict = "표본부족"
            elif bucket["label"] in top_labels:
                verdict = "🏆조회최적"
            elif bucket["median_views"] < best_median * 0.6:
                verdict = "⚠️조회데드존"
            else:
                verdict = "✅양호"
            ws.append([
                bucket["label"],
                bucket["count"],
                bucket["median_views"],
                bucket["avg_views"],
                bucket["median_er"],
                bucket["dead_rate"],
                bucket["view_share"],
                verdict,
            ])
            if verdict == "🏆조회최적":
                _fill_row(ws, row, len(headers), HIGHLIGHT_FILL)
            row += 1
        # 상충 캡션: ER 최고 구간과 조회 최고 구간의 실제 수치로 생성
        if eligible:
            er_best = max(eligible, key=lambda x: x["median_er"])
            view_best = max(eligible, key=lambda x: x["median_views"])
            if er_best["label"] != view_best["label"]:
                caption = (
                    f"글이 길수록 인게이지먼트율은 올라가지만 조회수는 급감합니다 "
                    f"({er_best['label']}: ER {er_best['median_er']}% / 조회 {er_best['median_views']}, "
                    f"{view_best['label']}: ER {view_best['median_er']}% / 조회 {view_best['median_views']}) "
                    f"— 조회 격차 {round(view_best['median_views'] / er_best['median_views'], 1)}배"
                    if er_best["median_views"] else ""
                )
            else:
                caption = f"{view_best['label']} 구간이 조회수와 인게이지먼트율 모두 최고입니다 (ER {view_best['median_er']}% / 조회 {view_best['median_views']})"
            if caption:
                ws.cell(row=row, column=1, value=caption)
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
                ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
                row += 1
    row += 1

    row = _set_section_title(ws, row, "2. 미디어타입별 상세 비교", end_col=9)
    headers = ["타입", "게시물수", "비중(%)", "평균조회", "평균좋아요", "평균답글", "평균리포스트", "평균인게이지먼트", "평균좋아요율(%)"]
    row = _set_headers(ws, row, headers)
    type_stats = {}
    for _, m in annotated:
        mt = m["media_type"]
        if mt not in type_stats:
            type_stats[mt] = []
        type_stats[mt].append(m)
    sorted_types = sorted(type_stats.items(), key=lambda x: _avg([m["engagement"] for m in x[1]]), reverse=True)
    for idx, (media_type, items) in enumerate(sorted_types, 1):
        ws.append([
            media_type,
            len(items),
            round(len(items) / len(annotated) * 100, 2) if annotated else 0,
            round(_avg([m["views"] for m in items])),
            round(_avg([m["likes"] for m in items]), 1),
            round(_avg([m["replies"] for m in items]), 1),
            round(_avg([m["reposts"] for m in items]), 1),
            round(_avg([m["engagement"] for m in items]), 1),
            round(_avg([m["like_rate"] for m in items]), 2),
        ])
        if idx == 1:
            _fill_row(ws, row, len(headers), HIGHLIGHT_FILL)
        row += 1
    row += 1

    row = _set_section_title(ws, row, "3. 캐러셀 vs 텍스트 직접 비교", end_col=4)
    headers = ["지표", "캐러셀", "텍스트", "배율(캐러셀/텍스트)"]
    row = _set_headers(ws, row, headers)
    carousel_items = [m for _, m in annotated if m["media_type"] == "CAROUSEL_ALBUM"]
    text_items = [m for _, m in annotated if m["media_type"] == "TEXT_POST"]

    def metric_value(items, field):
        if field == "count":
            return len(items)
        return _avg([m[field] for m in items]) if items else 0

    comparisons = [
        ("게시물수", metric_value(carousel_items, "count"), metric_value(text_items, "count")),
        ("평균조회", metric_value(carousel_items, "views"), metric_value(text_items, "views")),
        ("평균좋아요", metric_value(carousel_items, "likes"), metric_value(text_items, "likes")),
        ("평균답글", metric_value(carousel_items, "replies"), metric_value(text_items, "replies")),
        ("평균리포스트", metric_value(carousel_items, "reposts"), metric_value(text_items, "reposts")),
        ("평균인게이지먼트", metric_value(carousel_items, "engagement"), metric_value(text_items, "engagement")),
        ("평균좋아요율(%)", metric_value(carousel_items, "like_rate"), metric_value(text_items, "like_rate")),
    ]
    for label, carousel_value, text_value in comparisons:
        ratio = "-"
        if text_value:
            ratio = f"{(carousel_value / text_value):.2f}x"
        elif carousel_value:
            ratio = "∞"
        ws.append([
            label,
            round(carousel_value, 2) if label != "게시물수" else int(carousel_value),
            round(text_value, 2) if label != "게시물수" else int(text_value),
            ratio,
        ])
        row += 1
    row += 1

    row = _set_section_title(ws, row, "4. 시간대 × 요일 매트릭스", end_col=8)
    headers = ["시간", "월", "화", "수", "목", "금", "토", "일"]
    row = _set_headers(ws, row, headers)
    matrix = {(hour, day): [] for hour in range(24) for day in range(7)}
    for _, m in annotated:
        if not m["dt"]:
            continue
        matrix[(m["dt"].hour, m["dt"].weekday())].append(m["views"])
    matrix_values = []
    for hour in range(24):
        row_values = [f"{hour:02d}시"]
        for day in range(7):
            avg_views = round(_avg(matrix[(hour, day)])) if matrix[(hour, day)] else 0
            row_values.append(avg_views)
            matrix_values.append(avg_views)
        ws.append(row_values)
        row += 1
    non_zero_matrix = sorted(v for v in matrix_values if v > 0)
    percentile_80 = non_zero_matrix[int((len(non_zero_matrix) - 1) * 0.8)] if non_zero_matrix else 0
    matrix_start_row = row - 24
    for r in range(matrix_start_row, matrix_start_row + 24):
        for c in range(2, 9):
            value = ws.cell(row=r, column=c).value or 0
            if value >= percentile_80 and value > 0:
                ws.cell(row=r, column=c).fill = HIGHLIGHT_FILL
    row += 1

    row = _set_section_title(ws, row, "5. 토픽 태그별 성과", end_col=5)
    headers = ["토픽", "게시수", "중앙값조회", "중앙값ER(%)", "데드율(<500조회,%)"]
    row = _set_headers(ws, row, headers)
    # topic_tag는 토픽 없는 게시물에서 API 응답 키 자체가 빠진다
    topic_groups = {}
    for post, m in annotated:
        topic_groups.setdefault(post.get("topic_tag") or "", []).append(m)
    untagged = topic_groups.pop("", [])

    def topic_stats(label, items):
        views = [m["views"] for m in items]
        ers = [
            ((m["likes"] + m["replies"] + m["reposts"] + m["quotes"] + m["shares"]) / m["views"] * 100)
            if m["views"] > 0 else 0
            for m in items
        ]
        return [
            label,
            len(items),
            round(_median(views)),
            round(_median(ers), 2),
            round(len([v for v in views if v < 500]) / len(items) * 100, 2),
        ]

    if not topic_groups:
        ws.append(["데이터 없음 (topic_tag 미수집 — analyze.py 재실행 필요)", "-", "-", "-", "-"])
        row += 1
    else:
        small = [m for items in topic_groups.values() if len(items) < 5 for m in items]
        topic_rows = sorted(
            (topic_stats(label, items) for label, items in topic_groups.items() if len(items) >= 5),
            key=lambda r: -r[2],
        )
        if small:
            topic_rows.append(topic_stats("기타(n<5)", small))
        if untagged:
            topic_rows.append(topic_stats("토픽 없음", untagged))
        for topic_row in topic_rows:
            ws.append(topic_row)
            row += 1

    ws.freeze_panes = "A3"
    auto_width(ws)


def sheet_monthly_trend(wb, posts):
    ws = wb.create_sheet("월별 트렌드")
    active_posts = _active_posts(posts)
    annotated = [(p, _post_metrics(p)) for p in active_posts]
    monthly = {}
    for _, m in annotated:
        if not m["dt"]:
            continue
        month_key = m["dt"].strftime("%Y-%m")
        if month_key not in monthly:
            monthly[month_key] = []
        monthly[month_key].append(m)
    overall_avg_views = _avg([m["views"] for _, m in annotated]) if annotated else 0
    row = 1

    row = _set_section_title(ws, row, "1. 월별 성과 추이", end_col=10)
    headers = ["연월", "게시물수", "총조회", "평균조회", "총좋아요", "평균좋아요", "총답글", "인게이지먼트율(%)", "전월대비조회증감(%)", "전월대비좋아요증감(%)"]
    row = _set_headers(ws, row, headers)
    prev_views = None
    prev_likes = None
    month_rows = []
    for month in sorted(monthly):
        items = monthly[month]
        total_views = sum(m["views"] for m in items)
        total_likes = sum(m["likes"] for m in items)
        total_replies = sum(m["replies"] for m in items)
        total_engagement = sum(m["engagement"] for m in items)
        avg_views = round(total_views / len(items)) if items else 0
        avg_likes = round(total_likes / len(items), 1) if items else 0
        view_growth = "-" if prev_views in (None, 0) else round((total_views - prev_views) / prev_views * 100, 2)
        like_growth = "-" if prev_likes in (None, 0) else round((total_likes - prev_likes) / prev_likes * 100, 2)
        month_rows.append((month, avg_views))
        ws.append([
            month,
            len(items),
            total_views,
            avg_views,
            total_likes,
            avg_likes,
            total_replies,
            round((total_engagement / total_views * 100), 2) if total_views else 0,
            view_growth,
            like_growth,
        ])
        if avg_views > overall_avg_views:
            _fill_row(ws, row, len(headers), HIGHLIGHT_FILL)
        prev_views = total_views
        prev_likes = total_likes
        row += 1
    row += 1

    row = _set_section_title(ws, row, "2. 분기별 요약", end_col=5)
    headers = ["분기", "게시물수", "평균조회", "평균좋아요", "인게이지먼트율(%)"]
    row = _set_headers(ws, row, headers)
    quarterly = {}
    for _, m in annotated:
        if not m["dt"]:
            continue
        quarter = (m["dt"].month - 1) // 3 + 1
        key = f"{m['dt'].year} Q{quarter}"
        if key not in quarterly:
            quarterly[key] = []
        quarterly[key].append(m)
    for quarter_key in sorted(quarterly):
        items = quarterly[quarter_key]
        total_views = sum(m["views"] for m in items)
        total_engagement = sum(m["engagement"] for m in items)
        ws.append([
            quarter_key,
            len(items),
            round(_avg([m["views"] for m in items])),
            round(_avg([m["likes"] for m in items]), 1),
            round((total_engagement / total_views * 100), 2) if total_views else 0,
        ])
        row += 1
    row += 1

    row = _set_section_title(ws, row, "3. 성장 구간 분석 (시간 분위, 계정 독립)", end_col=5)
    headers = ["구간명", "기간", "게시물수", "평균조회", "평균좋아요", "인게이지먼트율(%)"]
    row = _set_headers(ws, row, headers)
    dated = sorted(((m["dt"], m) for _, m in annotated if m["dt"]), key=lambda x: x[0])
    phase_names = ["1분위(초기)", "2분위", "3분위", "4분위(최근)"]
    n = len(dated)
    if n == 0:
        ws.append(["데이터 없음", "-", 0, 0, 0, 0])
        row += 1
    else:
        for i, phase_name in enumerate(phase_names):
            start_i = i * n // 4
            end_i = (i + 1) * n // 4 if i < 3 else n
            chunk = dated[start_i:end_i]
            items = [m for _, m in chunk]
            if not items:
                continue
            label = f"{chunk[0][0].strftime('%Y-%m')}~{chunk[-1][0].strftime('%Y-%m')}"
            total_views = sum(m["views"] for m in items)
            total_engagement = sum(m["engagement"] for m in items)
            ws.append([
                phase_name,
                label,
                len(items),
                round(_avg([m["views"] for m in items])),
                round(_avg([m["likes"] for m in items]), 1),
                round((total_engagement / total_views * 100), 2) if total_views else 0,
            ])
            row += 1

    ws.freeze_panes = "A3"
    auto_width(ws)


def sheet_daily_views(wb, user_insights):
    ws = wb.create_sheet("일별 조회수 추이")
    series = []
    for d in user_insights.get("daily_views") or []:
        if not isinstance(d, dict):
            continue
        date = d.get("date")
        value = d.get("value", 0)
        if date and isinstance(value, (int, float)):
            series.append((str(date), int(value)))
    series.sort(key=lambda x: x[0])
    row = 1

    if not series:
        row = _set_section_title(ws, row, "일별 조회수 추이", end_col=4)
        ws.append(["데이터 없음 (일별 시계열 미수집)"])
        auto_width(ws)
        return

    values = [v for _, v in series]
    # 앞쪽 6일은 누적 구간 평균 (7일치가 아직 없음)
    moving_avg = [_avg(values[max(0, i - 6):i + 1]) for i in range(len(values))]

    row = _set_section_title(ws, row, "1. 최근 90일 일별 조회수", end_col=4)
    headers = ["날짜", "조회수", "7일이동평균", "전주동요일대비(%)"]
    row = _set_headers(ws, row, headers)
    start = max(0, len(series) - 90)
    mean_90 = _avg(values[start:])
    prev_ma = None
    for i in range(start, len(series)):
        date, value = series[i]
        wow = "-"
        if i >= 7 and values[i - 7] > 0:
            wow = round((value - values[i - 7]) / values[i - 7] * 100, 2)
        ws.append([date, value, round(moving_avg[i]), wow])
        if prev_ma is not None and (prev_ma < mean_90 <= moving_avg[i] or prev_ma > mean_90 >= moving_avg[i]):
            _fill_row(ws, row, len(headers), HIGHLIGHT_FILL)
        prev_ma = moving_avg[i]
        row += 1
    row += 1

    row = _set_section_title(ws, row, "2. 월별 실제 노출 조회수 (노출일 기준)", end_col=5)
    headers = ["연월", "일수", "총조회", "일평균조회", "전월대비(%)"]
    row = _set_headers(ws, row, headers)
    monthly = {}
    for date, value in series:
        monthly.setdefault(date[:7], []).append(value)
    prev_total = None
    for month in sorted(monthly):
        month_values = monthly[month]
        total = sum(month_values)
        # 부분 달(수집 시작월·진행중인 달)을 완전한 달과 비교하면 -90%대 가짜 급락이 나온다
        partial = len(month_values) < monthrange(int(month[:4]), int(month[5:7]))[1]
        mom = "-" if partial or prev_total in (None, 0) else round((total - prev_total) / prev_total * 100, 2)
        label = f"{month} (부분)" if partial else month
        ws.append([label, len(month_values), total, round(_avg(month_values)), mom])
        prev_total = None if partial else total
        row += 1
    row += 1

    row = _set_section_title(ws, row, "3. 요일별 노출량", end_col=4)
    headers = ["요일", "관측주수", "중앙값일조회", "평균일조회"]
    row = _set_headers(ws, row, headers)
    day_kr = {"Monday": "월", "Tuesday": "화", "Wednesday": "수", "Thursday": "목", "Friday": "금", "Saturday": "토", "Sunday": "일"}
    by_day = {name: [] for name in _WEEKDAY_NAMES}
    for date, value in series:
        try:
            by_day[_weekday_en(datetime.strptime(date, "%Y-%m-%d"))].append(value)
        except ValueError:
            continue
    medians = {name: _median(by_day[name]) for name in _WEEKDAY_NAMES}
    top2 = [name for name, _ in sorted(medians.items(), key=lambda x: x[1], reverse=True)[:2]]
    for name in _WEEKDAY_NAMES:
        day_values = by_day[name]
        ws.append([day_kr[name], len(day_values), round(medians[name]), round(_avg(day_values))])
        if day_values and name in top2:
            _fill_row(ws, row, len(headers), HIGHLIGHT_FILL)
        row += 1
    row += 1

    row = _set_section_title(ws, row, "4. 요약", end_col=2)
    row = _set_headers(ws, row, ["지표", "값"])
    last30 = sum(values[-30:])
    prev30 = sum(values[-60:-30])
    best = max(series, key=lambda x: x[1])
    worst = min(series, key=lambda x: x[1])
    for summary_row in [
        [f"{len(series)}일 총조회", f"{sum(values):,}"],
        ["최근 30일", f"{last30:,}"],
        ["직전 30일", f"{prev30:,}"],
        ["증감(%)", "-" if prev30 == 0 else round((last30 - prev30) / prev30 * 100, 2)],
        ["최고일", f"{best[0]} ({best[1]:,})"],
        ["최저일", f"{worst[0]} ({worst[1]:,})"],
    ]:
        ws.append(summary_row)
        row += 1
    row += 1
    ws.cell(row=row, column=1, value="API 기준 일 경계(UTC-07:00), 게시물 시각(JST)과 다름")

    ws.freeze_panes = "A3"
    auto_width(ws)


def sheet_longitudinal(wb, snapshots):
    ws = wb.create_sheet("스냅샷 성장추이")
    row = 1

    headers = ["스냅샷일시", "전체게시물", "원본게시물", "팔로워", "30일조회", "30일좋아요", "비고"]
    row = _set_section_title(ws, row, "1. 스냅샷 이력", end_col=len(headers))
    row = _set_headers(ws, row, headers)
    for s in snapshots:
        ws.append([
            s["label"],
            s["posts_total"],
            s["active"],
            s["followers"] if s["followers"] > 0 else "-",
            s["views_30d"],
            s["likes_30d"] if s["likes_30d"] > 0 else "-",
            " / ".join(s["flags"]) if s["flags"] else "정상",
        ])
        row += 1
    row += 1

    headers = ["구간", "일수", "원본게시물증가", "팔로워증가", "일평균팔로워", "원본1건당팔로워", "30일조회증감(%)"]
    row = _set_section_title(ws, row, "2. 구간별 성장률 (결측 스냅샷 구간은 팔로워 계산 제외)", end_col=len(headers))
    row = _set_headers(ws, row, headers)
    for prev, cur in zip(snapshots, snapshots[1:]):
        days = round((cur["dt"] - prev["dt"]).total_seconds() / 86400) if prev["dt"] and cur["dt"] else 0
        if days <= 0:
            continue  # 같은 시각 재수집본끼리의 0일 구간은 성장률이 아니다
        active_delta = cur["active"] - prev["active"]
        # 결측(followers<=0) 스냅샷을 그대로 빼면 가짜 급감이 나오므로 계산 자체를 안 한다
        if prev["followers"] > 0 and cur["followers"] > 0:
            follower_delta = cur["followers"] - prev["followers"]
            per_day = round(follower_delta / days, 2) if days > 0 else "-"
            per_post = round(follower_delta / active_delta, 2) if active_delta > 0 else "-"
        else:
            follower_delta = per_day = per_post = "-"
        view_change = "-"
        if prev["views_30d"] > 0:
            view_change = round((cur["views_30d"] - prev["views_30d"]) / prev["views_30d"] * 100, 2)
        ws.append([
            f"{prev['label']} → {cur['label']}",
            days, active_delta, follower_delta, per_day, per_post, view_change,
        ])
        if isinstance(follower_delta, int) and follower_delta > 0:
            _fill_row(ws, row, len(headers), HIGHLIGHT_FILL)
        row += 1
    row += 1

    headers = ["게시일후경과", "대상게시물수", "중앙값조회증가", "증가율(%)"]
    row = _set_section_title(ws, row, "3. 게시물 조회수 누적 곡선 (id 조인, n<20 구간은 '-')", end_col=len(headers))
    row = _set_headers(ws, row, headers)
    base = next((s for s in snapshots if s["by_id"]), None)
    latest = next((s for s in reversed(snapshots) if s["by_id"] and s is not base), None)
    buckets = [("0-7일", 0, 7), ("7-14일", 7, 14), ("14-30일", 14, 30),
               ("30-60일", 30, 60), ("60-90일", 60, 90), ("90일+", 90, float("inf"))]
    grouped = {label: [] for label, _, _ in buckets}
    if base and latest and base["dt"]:
        for post_id, (old_views, post_dt) in base["by_id"].items():
            new_entry = latest["by_id"].get(post_id)
            if not new_entry or not post_dt:
                continue
            age = (base["dt"] - post_dt).total_seconds() / 86400
            if age < 0:
                continue
            for label, low, high in buckets:
                if low <= age < high:
                    grouped[label].append((new_entry[0] - old_views, old_views))
                    break
    if any(grouped.values()):
        for label, _, _ in buckets:
            items = grouped[label]
            if len(items) >= 20:
                gain = round(_median([d for d, _ in items]))
                rate = round(_median([d / b * 100 for d, b in items if b > 0]), 2)
            else:
                gain = rate = "-"
            ws.append([label, len(items), gain, rate])
            row += 1
    else:
        ws.append(["데이터 없음", 0, "-", "-"])
        row += 1
    row += 1

    headers = ["스냅샷", "리포스트수", "비중(%)"]
    row = _set_section_title(ws, row, "4. 리포스트 비중 추이 (REPOST_FACADE)", end_col=len(headers))
    row = _set_headers(ws, row, headers)
    for s in snapshots:
        ws.append([
            s["label"],
            s["reposts"],
            round(s["reposts"] / s["posts_total"] * 100, 2) if s["posts_total"] else 0,
        ])
        row += 1

    ws.freeze_panes = "A3"
    auto_width(ws)


def sheet_growth_strategy(wb, posts, demographics, followers):
    ws = wb.create_sheet("10만 성장전략")
    active_posts = _active_posts(posts)
    annotated = [(p, _post_metrics(p)) for p in active_posts]
    total_views = sum(m["views"] for _, m in annotated)
    total_engagement = sum(m["engagement"] for _, m in annotated)
    engagement_rate = (total_engagement / total_views * 100) if total_views else 0
    viral_min = _viral_threshold(annotated)
    viral_ratio = (sum(1 for _, m in annotated if m["views"] >= viral_min) / len(annotated) * 100) if annotated else 0
    carousel_ratio = (sum(1 for _, m in annotated if m["media_type"] == "CAROUSEL_ALBUM") / len(annotated) * 100) if annotated else 0
    reply_post_ratio = (sum(1 for _, m in annotated if m["replies"] > 0) / len(annotated) * 100) if annotated else 0
    avg_like_rate = _avg([m["like_rate"] for _, m in annotated]) if annotated else 0
    hour_counter = Counter(m["dt"].hour for _, m in annotated if m["dt"])
    top_hours = sorted(hour_counter.items(), key=lambda x: x[1], reverse=True)[:2]
    crowded_hours = ",".join(f"{h:02d}시" for h, _ in top_hours) if top_hours else "데이터 부족"

    value_keywords = ["팁", "인사이트", "노하우", "정리", "가이드", "분석", "조언", "배운", "방법", "공식", "리뷰", "교육"]
    engagement_keywords = ["?", "어떻게", "어떤", "여러분", "생각", "의견", "알려", "추천", "맞나요", "인가요", "질문", "고견"]
    cta_keywords = ["링크", "프로필", "팔로우", "구독", "블로그", "뉴스레터", "DM", "문의", "보러", "공유"]
    mix_counter = Counter()
    for post, m in annotated:
        text = (post.get("text") or "").replace("\n", " ")
        if any(keyword in text for keyword in cta_keywords):
            mix_counter["소프트CTA"] += 1
        elif any(keyword in text for keyword in engagement_keywords):
            mix_counter["참여유도"] += 1
        elif m["media_type"] == "CAROUSEL_ALBUM" or any(keyword in text for keyword in value_keywords):
            mix_counter["가치콘텐츠"] += 1
        else:
            mix_counter["개인스토리"] += 1

    def judge_range(value, low, high=None, unit="%"):
        if high is None:
            if value >= low:
                return "🟢 적정"
            if value >= low * 0.7:
                return "🟡 근접"
            return "🔴 부족"
        if low <= value <= high:
            return "🟢 적정"
        margin = max((high - low) * 0.25, 0.5)
        if (low - margin) <= value <= (high + margin):
            return "🟡 근접"
        return "🔴 개선필요"

    deadzone_bucket = "201-300자 데드존"
    text_posts = [(p, m) for p, m in annotated if m["media_type"] == "TEXT_POST"]
    if text_posts:
        candidate_buckets = [
            ("~50자", [m["like_rate"] for p, m in text_posts if len((p.get("text") or "").replace("\n", " ").strip()) <= 50]),
            ("51-100자", [m["like_rate"] for p, m in text_posts if 51 <= len((p.get("text") or "").replace("\n", " ").strip()) <= 100]),
            ("101-200자", [m["like_rate"] for p, m in text_posts if 101 <= len((p.get("text") or "").replace("\n", " ").strip()) <= 200]),
            ("201-300자", [m["like_rate"] for p, m in text_posts if 201 <= len((p.get("text") or "").replace("\n", " ").strip()) <= 300]),
            ("300자+", [m["like_rate"] for p, m in text_posts if len((p.get("text") or "").replace("\n", " ").strip()) >= 301]),
        ]
        non_empty = [(label, _avg(rates)) for label, rates in candidate_buckets if rates]
        if non_empty:
            deadzone_bucket = f"{min(non_empty, key=lambda x: x[1])[0]} 데드존"

    row = 1
    row = _set_section_title(ws, row, "1. 현재 위치 진단", end_col=4)
    headers = ["지표", "현재값", "10만계정기준", "판정"]
    row = _set_headers(ws, row, headers)
    diagnosis_rows = [
        ["팔로워", f"{followers:,}" if followers else "-", "100,000", f"🔴 {followers / 100000 * 100:.1f}%" if followers else "-"],
        ["인게이지먼트율", f"{engagement_rate:.2f}%", "2-5%", judge_range(engagement_rate, 2, 5)],
        [f"바이럴비율({viral_min:,}+조회)", f"{viral_ratio:.2f}%", "5-10%", judge_range(viral_ratio, 5, 10)],
        ["캐러셀비중", f"{carousel_ratio:.2f}%", "10-20%", judge_range(carousel_ratio, 10, 20)],
        ["답글있는게시물", f"{reply_post_ratio:.2f}%", "70%+", judge_range(reply_post_ratio, 70)],
        ["평균좋아요율", f"{avg_like_rate:.2f}%", "3-5%", judge_range(avg_like_rate, 3, 5)],
    ]
    for values in diagnosis_rows:
        ws.append(values)
        row += 1
    row += 1

    carousel_items = [m for _, m in annotated if m["media_type"] == "CAROUSEL_ALBUM"]
    text_items = [m for _, m in annotated if m["media_type"] == "TEXT_POST"]
    if carousel_items and text_items:
        c_eng = _avg([m["engagement"] for m in carousel_items])
        t_eng = _avg([m["engagement"] for m in text_items])
        carousel_effect = f"캐러셀/텍스트 인게이지먼트 {c_eng / t_eng:.1f}배" if t_eng else "텍스트 대비 비교 불가"
    else:
        carousel_effect = "타입 비교 데이터 부족"

    hour_avg_views = {}
    for _, m in annotated:
        if not m["dt"]:
            continue
        hour_avg_views.setdefault(m["dt"].hour, []).append(m["views"])
    hour_avgs = {h: _avg(vs) for h, vs in hour_avg_views.items() if vs}
    if hour_avgs:
        best_h, best_v = max(hour_avgs.items(), key=lambda x: x[1])
        overall_v = _avg([m["views"] for _, m in annotated if m["dt"]]) or 1
        time_effect = f"최고 {best_h:02d}시 평균조회 {best_v / overall_v:.1f}배(전체대비)"
        time_target = f"{best_h:02d}시 집중"
    else:
        time_effect = "시간대 데이터 부족"
        time_target = "성과 상위 시간 분산"

    like_rates_by_len = []
    for p, m in annotated:
        if m["media_type"] != "TEXT_POST":
            continue
        text_len = len((p.get("text") or "").replace("\n", " ").strip())
        like_rates_by_len.append((text_len, m["like_rate"]))
    if like_rates_by_len:
        best_bucket_rates = []
        for low, high in [(0, 50), (51, 100), (101, 200), (201, 300), (301, 10**9)]:
            rates = [r for length, r in like_rates_by_len if low <= length <= high]
            if rates:
                best_bucket_rates.append(_avg(rates))
        if len(best_bucket_rates) >= 2:
            length_effect = f"길이구간 좋아요율 최대/최소 {max(best_bucket_rates) / max(min(best_bucket_rates), 0.01):.1f}배"
        else:
            length_effect = "길이 비교 데이터 부족"
    else:
        length_effect = "텍스트 게시물 부족"

    row = _set_section_title(ws, row, "2. 즉시 실행 TOP 5", end_col=5)
    headers = ["우선순위", "액션", "현재값", "목표값", "데이터 근거"]
    row = _set_headers(ws, row, headers)
    actions = [
        [1, "캐러셀 주2회", f"현재 {carousel_ratio:.1f}%", "10-15%", carousel_effect],
        [2, "게시시간 분산", f"{crowded_hours} 과밀", time_target, time_effect],
        [3, "최적 길이+질문마무리", deadzone_bucket, "성과 상위 길이대", length_effect],
        [4, "일일 사려깊은 답글 루틴", f"답글있는글 {reply_post_ratio:.1f}%", "70%+", "답글 비중↑ → 도달 가중 기대"],
        [5, "바이오 명확화", "프로필 점검", "가치제안 1문장", "전환 경로 명확화"],
    ]
    for values in actions:
        ws.append(values)
        if values[0] == 1:
            _fill_row(ws, row, len(headers), HIGHLIGHT_FILL)
        row += 1
    row += 1

    row = _set_section_title(ws, row, "3. 성장 로드맵 (현재 팔로워 기준 스케일)", end_col=5)
    headers = ["단계", "기간", "목표팔로워", "핵심액션", "KPI"]
    row = _set_headers(ws, row, headers)
    base = max(followers, 1)
    # Scale targets from current base toward 100k (or 3x if already large)
    t1 = max(int(base * 1.5), base + 500)
    t2 = max(int(base * 3), base + 2000)
    t3 = max(int(base * 8), base + 10000)
    t4 = max(100000, int(base * 15))
    roadmap = [
        ["Phase 1", "1-2개월", f"{t1:,}", "프로필최적화+캐러셀주2+시간조정", f"주 +{max(int(base * 0.05), 20):,} 팔로워"],
        ["Phase 2", "3-4개월", f"{t2:,}", "질문형마무리+답글루틴+참여팟", "인게이지먼트율2%+"],
        ["Phase 3", "5-8개월", f"{t3:,}", "콘텐츠기둥3개+바이럴공식반복+크로스채널", "월평균조회 상위화"],
        ["Phase 4", "9-18개월", f"{t4:,}", "권위구축+협업+데이터기반재활용", "바이럴비율5%+"],
    ]
    for values in roadmap:
        ws.append(values)
        row += 1
    row += 1

    row = _set_section_title(ws, row, "4. 콘텐츠 믹스 공식", end_col=5)
    headers = ["비중", "유형", "설명", "현재", "목표"]
    row = _set_headers(ws, row, headers)
    mix_rows = [
        ["40%", "가치콘텐츠", "팁/인사이트/교육", f"{(mix_counter['가치콘텐츠'] / len(annotated) * 100):.1f}%" if annotated else "0%", "40%"],
        ["30%", "개인스토리", "여정/비하인드", f"{(mix_counter['개인스토리'] / len(annotated) * 100):.1f}%" if annotated else "0%", "30%"],
        ["20%", "참여유도", "질문/투표/오픈스레드", f"{(mix_counter['참여유도'] / len(annotated) * 100):.1f}%" if annotated else "0%", "20%"],
        ["10%", "소프트CTA", "링크/전환", f"{(mix_counter['소프트CTA'] / len(annotated) * 100):.1f}%" if annotated else "0%", "10%"],
    ]
    for values in mix_rows:
        ws.append(values)
        row += 1

    ws.freeze_panes = "A3"
    auto_width(ws)


def parse_args(argv=None):
    import argparse

    parser = argparse.ArgumentParser(description="Threads 분석 JSON → Excel 리포트")
    parser.add_argument(
        "-i", "--input",
        default=None,
        help="analysis_*.json 경로 (미지정 시 output/ 최신 파일)",
    )
    parser.add_argument(
        "-o", "--output",
        default=None,
        help="출력 xlsx 경로 (미지정 시 output/threads_analysis_YYYYMMDD.xlsx)",
    )
    parser.add_argument(
        "--lang",
        default=None,
        choices=i18n.LANGS,
        help="리포트 언어 (미지정 시 THREADS_LANG, 없으면 ko)",
    )
    return parser.parse_args(argv)


def resolve_input_path(input_arg: str = None) -> str:
    import glob as _glob

    if input_arg:
        if not os.path.isfile(input_arg):
            print(f"[ERROR] 입력 파일 없음: {input_arg}")
            sys.exit(1)
        return input_arg

    output_dir = os.path.join(os.path.dirname(__file__), "output")
    pattern = os.path.join(output_dir, "analysis_*.json")
    json_files = sorted(_glob.glob(pattern))
    if not json_files:
        print("[ERROR] output/ 디렉토리에 analysis_*.json 파일이 없습니다.")
        print("  먼저 analyze.py를 실행해주세요.")
        sys.exit(1)

    # --max-posts 샘플 실행이 조용히 기본 리포트가 되는 사고를 막는다.
    loaded = []
    for path in json_files:
        try:
            loaded.append((path, load_data(path)))
        except (OSError, ValueError):
            continue
    if not loaded:
        return json_files[-1]
    marked = _mark_partials(loaded)
    full = [path for path, _, partial in marked if not partial]
    if full and full[-1] != json_files[-1]:
        skipped = sum(1 for _, _, partial in marked if partial)
        print(f"[알림] 부분 수집본 {skipped}개를 건너뛰고 완전 수집본을 사용합니다.")
        print("  특정 파일을 쓰려면 -i 로 지정하세요.")
    return full[-1] if full else json_files[-1]


# 부분 수집본 판정 기준. "partial": true 표식이 없는 옛 파일도 걸러야 한다.
# 전체 최대치와 비교하면 안 된다 — 계정이 성장하므로 과거 스냅샷은 당연히 적고,
# 1,822건짜리 4월 스냅샷이 2,292건 대비 부분 수집본으로 오판된다.
# 그 시점까지의 누적 최대치와 비교해야 "갑자기 급감했다"를 잡는다.
PARTIAL_RATIO = 0.9


def is_partial_snapshot(data, prior_max=0):
    """--max-posts 로 만든 부분 수집본인가. prior_max는 이 스냅샷 이전까지의 최대 게시물 수."""
    if data.get("partial"):
        return True
    posts = data.get("posts") or []
    return bool(prior_max) and len(posts) < prior_max * PARTIAL_RATIO


def _mark_partials(loaded):
    """[(path, data)]를 시간순으로 훑어 부분 수집본을 골라낸다. [(path, data, partial)] 반환."""
    def when(item):
        dt = parse_ts((item[1] or {}).get("analyzed_at"))
        return (dt is None, dt, item[0])

    out = []
    prior_max = 0
    for path, data in sorted(loaded, key=when):
        partial = is_partial_snapshot(data, prior_max)
        if not partial:
            prior_max = max(prior_max, len(data.get("posts") or []))
        out.append((path, data, partial))
    return out


def load_snapshots():
    """output/analysis_*.json 전체를 오래된 순으로 로드 (cross-snapshot 추이 전용).

    부분 수집본은 제외한다. 섞이면 게시물이 2,292 -> 40 으로 떨어진 것처럼 보여
    구간 성장률이 -1,918 같은 가짜 폭락을 낸다.
    """
    import glob as _glob

    output_dir = os.path.join(os.path.dirname(__file__), "output")
    paths = sorted(_glob.glob(os.path.join(output_dir, "analysis_*.json")))

    loaded = []
    for path in paths:
        try:
            loaded.append((path, load_data(path)))
        except (OSError, ValueError):
            continue  # 수집 중이거나 깨진 파일은 건너뜀
    snapshots = []
    for path, data, partial in _mark_partials(loaded):
        if partial:
            continue
        posts = data.get("posts") or []
        user_insights = data.get("user_insights") or {}
        dt = parse_ts(data.get("analyzed_at"))
        followers = user_insights.get("followers_count") or 0
        likes_30d = user_insights.get("30d_likes") or 0
        flags = []
        if followers <= 0:
            flags.append("팔로워 데이터 결측")
        if likes_30d <= 0:
            flags.append("30일좋아요 데이터 결측")
        snapshots.append({
            "label": dt.strftime("%Y-%m-%d %H:%M") if dt else os.path.basename(path),
            "dt": dt,
            "posts_total": len(posts),
            "reposts": sum(1 for p in posts if p.get("media_type") == "REPOST_FACADE"),
            "active": len(_active_posts(posts)),
            "followers": followers,
            "views_30d": user_insights.get("30d_views") or 0,
            "likes_30d": likes_30d,
            "flags": flags,
            # id → (조회수, 게시일). 리포스트는 insights가 항상 비어 있어 제외
            "by_id": {
                p["id"]: ((p.get("insights") or {}).get("views", 0), parse_ts(p.get("timestamp")))
                for p in _active_posts(posts) if p.get("id")
            },
        })
    return snapshots


def build_workbook(data: dict):
    posts = data.get("posts", [])
    user_insights = data.get("user_insights", {})
    demographics = data.get("follower_demographics", {})
    followers = user_insights.get("followers_count", 0)
    username = data.get("profile", {}).get("username", "unknown")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    sheet_all_posts(wb, posts)
    sheet_ranking(wb, posts, followers)
    sheet_time_analysis(wb, posts)
    sheet_demographics(wb, demographics, followers)
    sheet_insights_report(wb, posts, user_insights, followers, username, demographics)
    sheet_viral_analysis(wb, posts)
    sheet_like_rate_analysis(wb, posts)
    sheet_content_optimization(wb, posts)
    sheet_monthly_trend(wb, posts)
    sheet_daily_views(wb, user_insights)
    snapshots = load_snapshots()
    if len(snapshots) >= 2:
        sheet_longitudinal(wb, snapshots)
    sheet_growth_strategy(wb, posts, demographics, followers)
    return wb


def main(argv=None):
    args = parse_args(argv)
    input_path = resolve_input_path(args.input)

    output_dir = os.path.join(os.path.dirname(__file__), "output")
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now(JST).strftime("%Y%m%d")
    output_path = args.output or os.path.join(
        output_dir, f"threads_analysis_{timestamp}.xlsx"
    )

    lang = i18n.resolve_lang(args.lang)
    print(f"입력: {input_path}")
    if lang != "ko":
        print(f"언어: {lang}")
    data = load_data(input_path)
    wb = build_workbook(data)
    if i18n.translate_workbook(wb, lang) is not None and lang != "ko":
        # 열 너비는 시트를 만들 때 한국어 기준으로 계산됐다. 영문 헤더가 더 길어 잘리므로 다시 잰다.
        for ws in wb.worksheets:
            auto_width(ws)
    wb.save(output_path)
    print(f"Excel 저장 완료: {output_path}")
    print(f"시트: {wb.sheetnames}")
    followers_count = data.get("user_insights", {}).get("followers_count")
    print(f"게시물: {len(data.get('posts', []))}개 / 팔로워: {f'{followers_count:,}' if followers_count else '-'}")


if __name__ == "__main__":
    main()
